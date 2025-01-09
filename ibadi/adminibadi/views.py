from django.shortcuts import render,redirect,get_object_or_404,HttpResponse
from django.urls import reverse
from decimal import Decimal
from .forms import aloginForm, CouponForm
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth import authenticate,login,logout
from useribadi.models import User, Order,OrderItem,Wallet
from django.db.models import Q
from adminibadi.models import Category,Product,ProductImage,ProductVariants,Coupon
from django.views.decorators.cache import cache_control
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
from django.db.models import Sum,Count
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from django.utils.timezone import datetime





# Create your views here.

@cache_control(no_store=True, must_revalidate=True, no_cache=True)
def adminLogin(request):
    if request.user.is_authenticated:
         if hasattr(request.user, 'is_admin') and request.user.is_admin:
            return redirect('adminHome') 

    if request.method == 'POST':
        form = aloginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            user = authenticate(request, email=email, password=password)
            if user is not None:
                if user.is_admin:
                    login(request,user)
                    messages.success(request,'Admin LoggedIn')
                    return redirect('adminHome')
                else:
                    messages.error(request, 'Invalid email or Password')
            else:
                messages.error(request,'Invalid email or Password')
            
    else:
        form = aloginForm()

    return render(request,'adminibadi/alogin.html', {'form' : form})


@cache_control(no_store=True, must_revalidate=True, no_cache=True)
@login_required
def adminHome(request):
    if not request.user.is_authenticated:
            return redirect('adminLogin')
    return render(request,'adminibadi/adminHome.html')


def adminLogout(request):
    logout(request)
    messages.success(request,'Admin logged out successfully')
    return redirect('adminLogin')


def customers(request):
    users = User.objects.filter(is_admin=False).order_by('-id')
    return render(request,'adminibadi/customers.html',{'users' : users })
    


def blockUser(request,id):
     user = get_object_or_404(User,id=id)
     if request.user.is_authenticated:
            user.is_active = False
            user.save()
            messages.error(request,f'{user.full_name} has been blocked!')
     return redirect('customers')

def unblockUser(request,id):
    user = get_object_or_404(User,id=id)
    if request.user.is_authenticated:
        user.is_active = True
        user.save()
        messages.success(request,f'{user.full_name} has been unblocked!')
    return redirect('customers')




def category(request):
    categories = Category.objects.filter(is_deleted = False)
    return render(request,'adminibadi/categories.html', {'categories' : categories})



def addCategory(request):
    if request.method == 'POST':
        category_name = request.POST.get('category_name')
        description = request.POST.get('description')

        if not category_name:
            messages.error(request,'Category name is required! ')
            return render(request,'adminibadi/addCategory.html')
        

        if not all(char.isalpha() or char.isspace() for char in category_name):
            messages.error(request,'Category name must be contain alphabetic characters')
            return render(request,'adminibadi/addCategory.html')
        

        if Category.objects.filter(category_name__iexact=category_name).exists():
            messages.error(request,f'Category {category_name} is already exists')
            return render(request,'adminibadi/addCategory.html')
        
        
        Category.objects.create(category_name=category_name, description= description, is_active=True, is_deleted=False)
        messages.success(request,f'Category {category_name} is addedd successfully')
        return redirect('addCategory')
    
    return render(request,'adminibadi/addCategory.html')



def editCategory(request,category_id):
    category = get_object_or_404(Category, pk=category_id, is_deleted=False)

    if request.method == 'POST':
        if 'update' in request.POST:
            category_name = request.POST.get('category_name')
            description = request.POST.get('description')

            if not category_name:
                messages.error(request,'Category name is required')
                return redirect('editCategory', category_id=category_id)

            if not all(char.isalpha() or char.isspace() for char in category_name):
                messages.error(request,'Category name must be contain alphabetic characters')
                return redirect('editCategory', category_id=category_id)
            
            if Category.objects.filter(category_name__iexact=category_name).exists():
                messages.error(request,f'Category {category_name} is already exists')
                return redirect('editCategory', category_id=category_id)
        
            else:

                category.category_name = category_name
                category.description = description
                category.save()
                messages.success(request,'Category updated Successfully')
                return redirect('category')
        
        elif 'delete' in request.POST:
            category.is_deleted = True
            category.is_active = False
            category.save()
            messages.error(request,'Category deleted ')
            return redirect('category')
 
    return render(request,'adminibadi/editCategory.html',{'category':category})


def categoryStatus(request, category_id):
    category = get_object_or_404(Category, category_id=category_id, is_deleted=False)
    category.is_active = not category.is_active
    category.save()

    action = 'listed' if category.is_active else 'unlisted'
    messages.success(request,f'Category {category.category_name} has been {action}')
    return redirect('category')


def productsList(request):
    products = Product.objects.filter(is_deleted=False).order_by('-created_at').prefetch_related('product_images').select_related('category','variant')

    for product in products:
        product.effectiveOfferPercentage = max(
            product.product_offer_percentage or 0,
            product.category.category_offer_percentage or 0
            )
    return render(request,'adminibadi/products.html',{'products' : products})





def addProduct(request):
    categories = Category.objects.filter(is_active=True)
    variants = ProductVariants.objects.filter(is_active=True)

    if request.method == 'POST':
        product_name = request.POST.get('product_name')
        category_id = request.POST.get('category')
        variant_id = request.POST.get('variant')
        description = request.POST.get('description')
        quantity = request.POST.get('quantity')
        regular_price = request.POST.get('regular_price')
        selling_price = request.POST.get('selling_price')

        if not all([ product_name, category_id, variant_id,description, quantity, regular_price, selling_price]):
            messages.error(request,'Fields is required') 
            return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})
        
        if not all (char.isalpha() or char.isspace() for char in product_name):
            messages.error(request,'Product name must contain only alphabetic characters')
            return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})
        
        if not quantity.isdigit() or int(quantity) < 0:
            messages.error(request,'Quantity must be positive number')
            return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})
        
        try:
            regular_price = float(regular_price)
            selling_price = float(selling_price)

            if regular_price <= 0 or selling_price <= 0:
                raise ValueError
            if selling_price > regular_price:
                messages.error(request, 'Selling price must be equal or lessthan the regualr price')
                return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})
            
        except ValueError:
            messages.error(request,'Price must be positive number')
            return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})
        
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        image3 = request.FILES.get('image3')
        if not all([image1, image2, image3]):
            messages.error(request,'Please upload three images')
            return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})
            

        try:
            category = Category.objects.get(category_id=category_id)
            variant = ProductVariants.objects.get(variant_id=variant_id)
        except Category.DoesNotExist:
            messages.error(request, 'Invalid category selected')
            return render(request,'adminibadi/addProduct.html', {'categories' : categories, 'variants':variants})
        except ProductVariants.DoesNotExist:
            messages.error(request,'Invalid variant selection')
            return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})

        product = Product(product_name=product_name, category=category,variant=variant, description=description, regular_price=regular_price, selling_price=selling_price, quantity=quantity)
        product.save()

        ProductImage.objects.create(product=product, images=image1, is_main=True)
        ProductImage.objects.create(product=product, images=image2)
        ProductImage.objects.create(product=product, images=image3)
        messages.success(request, 'Product Successfully Added!')
        return redirect('productsList')
    return render(request,'adminibadi/addProduct.html',{'categories' : categories, 'variants':variants})



def editProduct(request,product_id):
    product = get_object_or_404(Product, product_id=product_id)
    categories = Category.objects.filter(is_active=True)
    variants = ProductVariants.objects.filter(is_active=True)
    images = ProductImage.objects.filter(product=product)

    imageMap = {f'image{i+1}': img for i,img in enumerate(images)}
    imageFileName = {key: image.images.name.split('/')[-1] for key, image in imageMap.items()}

    if request.method == 'POST':
        if 'update' in request.POST:
            product_name = request.POST.get('product_name')
            category_id = request.POST.get('category')
            description = request.POST.get('description')
            variant_id = request.POST.get('variant')
            quantity = request.POST.get('quantity')
            regular_price = request.POST.get('regular_price')
            selling_price = request.POST.get('selling_price')

            if not all (char.isalpha() or char.isspace() for char in product_name):
                messages.error(request,'Product name must contain only alphabetic characters')
                return redirect('editProduct', product_id=product_id)
            

            if not quantity.isdigit() or int(quantity) < 0:
                messages.error(request,'Quantity must be a positive number')
                return redirect('editProduct', product_id=product_id)
            
            try:
                regular_price = float(regular_price)
                selling_price = float(selling_price)
                if regular_price <= 0 or selling_price <= 0 :
                    raise ValueError
                if selling_price > regular_price:
                    messages.error(request,'Selling price must be equal or less than regular price')
                    return redirect('editProduct', product_id=product_id)
            except ValueError:
                messages.error(request,'Price must be positive number')
                return redirect('editProduct', product_id=product_id)
            

            try: 
                category = Category.objects.get(category_id=category_id)
                variant = ProductVariants.objects.get(variant_id=variant_id)
            except Category.DoesNotExist():
                messages.error(request,'Invalid category selected')
                return redirect('editProduct', product_id=product_id)

            product.product_name = product_name
            product.category = category
            product.variant = variant
            product.description = description
            product.quantity = quantity
            product.regular_price = regular_price
            product.selling_price = selling_price
            product.save()

            image1 = request.FILES.get('image1')
            image2 = request.FILES.get('image2')
            image3 = request.FILES.get('image3')

            if not images.exists() and not any([image1, image2, image3]):
                messages.error(request,'Images are required')
                return redirect('editProduct', product_id=product_id)

            if image1:
                oldImage1 = ProductImage.objects.filter(product=product,is_main=True).first()
                if oldImage1:
                    oldImage1.delete()
                ProductImage.objects.create(product=product,is_main=True, images=image1)

            if image2:
                oldImage2 = ProductImage.objects.filter(product=product,is_main=False).first()
                if oldImage2:
                    oldImage2.delete()
                ProductImage.objects.create(product=product, is_main=False,images=image2)

            if image3:
                oldImage3 = ProductImage.objects.filter(product=product,is_main=False).last()
                if oldImage3:
                    oldImage3.delete()
                ProductImage.objects.create(product=product,is_main=False,images=image3)
        
            messages.success(request,'Product Uploaded Successfully! ')
            return redirect('productsList')
        
        elif 'delete' in request.POST:
            product.is_active = False
            product.is_deleted = True
            product.save()

            messages.error(request,'Product deleted Successfully')
            return redirect('productsList')

    return render(request,'adminibadi/editProduct.html',{
        'product':product, 
        'categories':categories,
        'variants' : variants,
        'imageFileName' : imageFileName,
        })

def productStatus(request,product_id):
    product = get_object_or_404(Product, product_id=product_id, is_deleted=False)

    product.is_active = not product.is_active
    product.save()
    status = 'listed' if product.is_active else 'unlisted'
    messages.success(request, f'Product "{product.product_name}" { status }')
    return redirect('productsList')



def ordersList(request):
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        newStatus = request.POST.get('order_status')

        try:
            order = Order.objects.get(order_id=order_id)
            if newStatus in dict(order.order_status_choices):
                order.order_status = newStatus
                if order.order_status == 'DELIVERED':
                    order.payment_status = 'PAID'
                order.save()
                print(newStatus)
                print(order.payment_status)
                return redirect('ordersList')  
        except Order.DoesNotExist:
            return HttpResponse('Order not found',status=404)
    orders = Order.objects.select_related('user').all().order_by('-order_id')

    paginator = Paginator(orders,10)
    page_number = request.GET.get('page',1)
    page_obj = paginator.get_page(page_number)
    return render(request,'adminibadi/orders.html',{'orders':page_obj})



def orderDetails(request,order_id):
    order = get_object_or_404(Order,order_id=order_id)
    print(order.order_id)
    print(order.user.full_name)

    items = OrderItem.objects.filter(order=order)
    for item in items:
        item.main_image = item.product.product_images.filter(is_main=True).first()
        item.total_price = item.quantity * item.price
    return render(request,'adminibadi/orderDetails.html',{'order':order, 'items':items})



def acceptReturn(request,order_id):
    order = get_object_or_404(Order, order_id=order_id)
    if order.order_status == 'Return Requested':
        order.order_status = 'Returned'
        order.payment_status = 'REFUNDED'
        order.save()

        refundableAmount = order.final_amount - order.delivery_charge

        latestWalletEntry = Wallet.objects.filter(user=order.user).order_by('-created_at').first()
        currentBalance = latestWalletEntry.current_balance if latestWalletEntry else Decimal(0)
        updatedBalance = currentBalance + refundableAmount

        Wallet.objects.create(
            user = order.user,
            transaction_type = 'CREDITED',
            order = order,
            amount = refundableAmount,
            current_balance = updatedBalance,
            reason = f'Refund for returned order {order.order_id}'
        )
        
        orderItems = OrderItem.objects.filter(order=order)
        for item in orderItems:
            product = item.product
            product.quantity += item.quantity
            product.save()

        messages.success(request,'Return request accepted and amount refunded to wallet.')
    else:
        messages.error(request, 'Invalid return request status.')
    return redirect('orderDetails',order_id)




def rejectReturn(request,order_id):
    order = get_object_or_404(Order,order_id=order_id)
    if order.order_status == 'Return Requested':
        order.order_status = 'Return Rejected'
        order.save()
        messages.error(request, f'Return request rejected for order {order.order_id}')
    else:
        messages.error(request,'Invalid return request status')
    return redirect('orderDetails',order_id)




def addProductOffer(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, product_id=product_id)
        productOffer = Decimal(request.POST.get('offer_percentage',0))
        if 0 <= productOffer <= 100:
            categoryOffer = product.category.category_offer_percentage or Decimal(0)
            print(categoryOffer)
            print(productOffer)

            if productOffer > categoryOffer:
                applicableOffer = productOffer
            else:
                applicableOffer = categoryOffer

            print(applicableOffer)
            discountAmount = (product.regular_price * applicableOffer) / 100
            product.selling_price = product.regular_price - discountAmount
            product.product_offer_percentage = productOffer
            product.save()
            messages.success(request, f'Offer applied to {product.product_name}')
        else:
            messages.error(request, 'Please provide valid offer percentage')
    return redirect(reverse('productsList'))



def addCategoryOffer(request, category_id):
    if request.method == 'POST':
        category = get_object_or_404(Category,category_id=category_id)
        try:
            categoryOffer = Decimal(request.POST.get('category_offer', 0))
        except:
            messages.error(request, 'Please provide valid offer percentage')
            return redirect(reverse('category'))
        
        
        if 0 <= categoryOffer <= 100:
            category.category_offer_percentage = categoryOffer
            category.save()

            products = Product.objects.filter(category=category, is_active=True)
            for product in products:
                productOffer = product.product_offer_percentage or Decimal(0)
                print(categoryOffer)
                print(productOffer)

                if categoryOffer > productOffer:
                    applicableOffer = categoryOffer
                else:
                    applicableOffer = productOffer
                
                print(applicableOffer)
                discountAmount = (product.regular_price * applicableOffer) / 100
                product.selling_price = product.regular_price - discountAmount
                # product.product_offer_percentage = categoryOffer
                product.save()
            messages.success(request, f'Offer applied to catgory {category.category_name}')
        else:
            messages.error(request,'Pleases provide valid offer percentage')
    return redirect(reverse('category'))



def coupons(request):
    coupons = Coupon.objects.filter(is_active = True)
    return render(request, 'adminibadi/coupons.html',{'coupons':coupons})


def addCoupon(request):
    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            coupon_code = form.cleaned_data['coupon_code']
            coupon_name = form.cleaned_data['coupon_name']
            discount_percentage = form.cleaned_data['discount_percentage']
            minimum_purchase = form.cleaned_data['minimum_purchase']
            maximum_discount = form.cleaned_data['maximum_discount']
            expiry_date = form.cleaned_data['expiry_date']
            if maximum_discount > minimum_purchase:
                messages.error(request,'Maximum discount should be less than minimum purchase')
            else:
                Coupon.objects.create(
                    coupon_code = coupon_code,
                    coupon_name = coupon_name,
                    discount_percentage = discount_percentage,
                    minimum_purchase = minimum_purchase,
                    maximum_discount = maximum_discount,
                    expiry_date = expiry_date
            )
            messages.success(request,'Coupon addedd succeffully')
            return redirect('coupons')

        else:
            messages.error(request, 'Please correct the error below')
    else:
        form = CouponForm()
    return render(request, 'adminibadi/addCoupon.html', {'form':form})




def editCoupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, coupon_id = coupon_id)
    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            coupon.coupon_code = form.cleaned_data['coupon_code']
            coupon.coupon_name = form.cleaned_data['coupon_name']
            coupon.discount_percentage = form.cleaned_data['discount_percentage']
            coupon.minimum_purchase = form.cleaned_data['minimum_purchase']
            coupon.maximum_discount = form.cleaned_data['maximum_discount']
            coupon.expiry_date = form.cleaned_data['expiry_date']

            coupon.save()
            messages.success(request, 'Coupon updated successfully')
            return redirect('coupons')
        else:
            messages.error(request, 'Please give the valid inputs')
    else:
        form = CouponForm(initial={
            'coupon_code': coupon.coupon_code,
            'coupon_name':coupon.coupon_name,
            'discount_percentage':coupon.discount_percentage,
            'minimum_purchase':coupon.minimum_purchase,
            'maximum_discount':coupon.maximum_discount,
            'expiry_date':coupon.expiry_date
        })
    return render(request, 'adminibadi/editCoupon.html', {'form': form, 'coupon': coupon})



def deleteCoupon(request, coupon_id):
    if request.method == 'POST':
        coupon = get_object_or_404(Coupon, coupon_id=coupon_id)
        coupon.delete()
        messages.error(request, 'Coopen deleted successfully')
        return redirect('coupons')
    return redirect('coupons')



def salesReport(request):
    filterType = request.GET.get('filter_type', 'daily')
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    
    orders = Order.objects.filter(payment_status='PAID').order_by('-order_id')

    if filterType == 'custom' and startDate and endDate:
        orders=orders.filter(order_at__range=[startDate,endDate])
    elif filterType == 'daily':
        today = datetime.today()
        orders = orders.filter(order_at__year=today.year, order_at__month=today.month, order_at__day=today.day)
    elif filterType == 'weekly':
        startOfWeek = datetime.today() - timedelta(days=7)
        orders = orders.filter(order_at__gte=startOfWeek)
    elif  filterType == 'monthly':
        stratOfMonth = datetime.today() - timedelta(days=30)
        orders = orders.filter(order_at__gte=stratOfMonth)

    paginator = Paginator(orders, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    overallSalesCount = orders.count()
    overallOrderAmount = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    overallDiscount = orders.aggregate(Sum('discount_amount'))['discount_amount__sum'] or 0

    if 'download_pdf' in request.GET:
        return pdfReport(orders)
    if 'download_excel' in request.GET:
        return excelReport(orders)
    
    return render(request, 'adminibadi/salesReport.html',{
        'page_obj' : page_obj,
        'overall_sales_count' : overallSalesCount,
        'overall_order_amount' : overallOrderAmount,
        'overall_discount' : overallDiscount,
        'filter_type' : filterType,
        'start_date' : startDate,
        'end_date' :endDate,
    })



def pdfReport(orders):
    html = render_to_string('adminibadi/salesReportPdf.html',{'orders':orders})

    buffer = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode('UTF-8')), buffer)

    if not pdf.err:
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="sales_report.pdf"'
        return response
    else:
        return HttpResponse('Error generating PDF', status=500)
    



def excelReport(orders):
    html = render_to_string('adminibadi/salesReportExcel.html',{'orders':orders})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Report"

    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html,'html.parser')
    table = soup.find('table')

    headerRow = table.find('thead').find_all('th')
    for col_num, header in enumerate(headerRow, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header.text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    bodyRows = table.find('tbody').find_all('tr')
    for row_num, row in enumerate(bodyRows, start=2):
        cols = row.find_all('td')
        for col_num, col in enumerate(cols, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = col.text
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length,len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    workbook.save(response)
    return response












                



